from tkinter import Tk
from tkinter.filedialog import askopenfilename
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet import dimensions
from openpyxl.worksheet.worksheet import Worksheet

from typing import TypeAlias, Literal

from spl_widgets.autoscorer.tokenize_to_ipa import *
from spl_widgets.util.color_util import to_rich_text

AutoscorerTokens: TypeAlias =  list[ tuple[str, int, bool|None] ]
ScoringMode: TypeAlias = Literal['preferred-transcription', 'best-match']

# this version of the autoscorer uses a token list instead of a string, allowing
# for the treatment of 2+ wide characters (e.g. diphthongs) as single tokens
def score_transcription(
    sentence_ipa: list[str],
    transcription_ipa: list[str],
    _prev_scored: AutoscorerTokens = ...,
    _prev_transcription_idx: int=0
    ) -> tuple[int, AutoscorerTokens]:
    """
    Scores the inputted transcription IPA based on the passed target sentence IPA using recursion

    Parameters
    ----------
        @param sentence_ipa ( list[str] ): the input IPA (tokenized) of the target sentence
        @param transcription_ipa ( list[str] ): the input IPA (tokenized) of the subject's transcription

    ### Internal Parameters (Not to be passed by user)

        @param _prev_scored ( list[tuple[str,int,bool|None]] ):
        the tokens already scored

        @param _prev_transcription_idx ( int ):
            the current absolute index of the start of the passed transcription IPA. Added to a token's  relative
             transcription IPA index when it is appended to _prev_scored (transcription_ipa is consumed during recursion)

    Returns
    -------
        @returns score ( int ): the maximum allowable score for the subject's transcription
        @returns prev_scored ( list[tuple[str, int, bool|None]] ): the scored tokens

    ---

    Summary
    -------

    This function does the main legwork of the module, taking in the target and transcribed sentences as IPA
    and returning the maximum allowable score by recursively parsing them, and storing each processed token
    in the transcribed IPA (as well as omitted tokens from the target IPA) in a list for formatting by output_scoring()

    ### NB: Information about the contents of _prev_scored:

    _prev_scored is a list of tokens (tuples) of the form ( str, int, bool | None )
    Each of these contains information that tells the program how to score and output the token (see below):

    - str: the token text
    - int: the position in the transcription IPA
    - bool|None: whether the token text is valid -> True: valid (Score 1), False: extraneous, None: omitted

    """

    def curr_score():
        return sum((n[2] is True) for n in _prev_scored)

    def last_token_idx():
        return max(n[1] for n in _prev_scored)

    # reached end of sentence IPA, so we mark any further transcription IPA tokens as extraneous and return
    if sentence_ipa == []:
        lti = last_token_idx()
        for i, token in enumerate(transcription_ipa, start=1):
            _prev_scored.append((token,lti+i,False))

        return (curr_score(), _prev_scored)

    # if we are at the top level of our function, initialize _prev_scored
    if _prev_transcription_idx==0:
        _prev_scored = []

    sentence_ipa_idx = 0

    for i, token in enumerate(transcription_ipa):
        absolute_idx = i+_prev_transcription_idx
        if not token.isalpha():                      # non-IPA (space) token
            continue

        # get next IPA char from sentence, skipping spaces etc. 
        while not sentence_ipa[sentence_ipa_idx].isalpha():
            sentence_ipa_idx += 1

        if token == sentence_ipa[sentence_ipa_idx]:
            # The character is both valid (exists in sentence IPA) and is the character we are currently
            # looking for (first IPA character in sentence IPA)

            # so take it no matter what (shouldn't affect score but please fact-check me I'm not sure)
            return score_transcription(
                sentence_ipa[sentence_ipa_idx + 1:],
                transcription_ipa[i+1:],
                _prev_scored+[(token,absolute_idx,True)],
                absolute_idx+1
            )

        elif token in sentence_ipa:
            # the character could be valid but it is ahead of the current position in the sentence, so we optimize
            # score by running two scoring passes on the remaining IPA – one where we consider the character
            # extraneous and stay at the same point in the source (sentence) IPA, and one where we consider
            # it valid and jump forward in the source IPA, considering all intervening characters extraneous.
            # We then return the maximum score between these two runs as our optimized score

            # get the potential jump-ahead index of the character in the sentence IPA
            new_sentence_ipa_idx = sentence_ipa.index(token)

            # make a token we can append to _prev_scored if we decide to omit these characters
            tokens_between = "".join(                                    # get the omitted tokens
                sentence_ipa[sentence_ipa_idx:new_sentence_ipa_idx]
            )

            # get the index of the first omitted char in this group
            omitted_idx = 0                 # prevent the program breaking if the omitted character is the very first of the transcription
            if len(_prev_scored) > 0:
                omitted_idx = max(n[1] for n in _prev_scored)+1
            omitted_chars = [(tokens_between.strip(), omitted_idx, None)]            # create the token

            return max(
                # take this char as the next valid one and jump forward to it in the sentence IPA
                score_transcription(
                    sentence_ipa[new_sentence_ipa_idx+1:],
                    transcription_ipa[i+1:],
                    _prev_scored + omitted_chars + [(token,absolute_idx,True)],
                    absolute_idx+1
                ),
                # take this char as extraneous and move to the next character without moving forward in the sentence IPA
                score_transcription(
                    sentence_ipa,
                    transcription_ipa[i+1:],
                    _prev_scored + [(token,absolute_idx,False)],
                    absolute_idx+1
                ),
                key = lambda n: n[0]    # get max by score
            )

        # character is not found in the sentence IPA, and thus could not
        # possibly be valid – so, we mark it as extraneous and move on
        else:
            _prev_scored.append((token,absolute_idx,False))

    # reached end of transcription IPA, add any remaining sentence IPA tokens as omitted and return
    _prev_scored.append(("".join(sentence_ipa), last_token_idx()+1, None))
    return (curr_score(), _prev_scored)

# this is probably not incredibly elegant (the colorizing part) but it does its job, and it shouldn't ever need to
# be changed too much functionally (except to add more features) because the data structure it gets from score() will stay the same
def output_scoring(
    sentence_ipa: list[str],
    transcription_ipa: list[str],
    results: tuple[ int, AutoscorerTokens ]
    ) -> tuple[int,int,str]:
    """
    Takes the results of score_transcription() and outputs them as formatted strings
    containing the autoscorer's evaluation and score for the transcription

    Parameters
    ----------
        @param sentence_ipa ( list[str] ): the input IPA (tokenized) of the target sentence
        @param transcription_ipa ( str ): the input IPA (tokenized) of the subject's transcription
        @param results ( int, list[tuple[str, int, bool|None]] ): results returned from score_transcription (see function docstring for details)

    Returns
    -------
        @returns score ( int ): the autoscorer's score for the transcription
        @returns best_poss_score ( int ): the best possible score for the target sentence
        @returns evaluation ( str ): the autoscorer's evaluation of the transcription's IPA, to be passed to to_rich_text

    """

    # color codes for to_rich_text()
    RED = "\R"
    GREEN = "\G"
    YELLOW = "\Y"
    CLEARSTYLE = "\C"

    score, prev_scored = results

    best_poss_score = sum( map(str.isalpha, sentence_ipa) )
    transcription_chars = [*transcription_ipa]                  # split the string

    # separate tokens into included and omitted for the two formatting passes
    included_tokens = [*filter(lambda n: n[2] is not None, prev_scored)]
    omitted_tokens = [*filter(lambda n: n[2] is None, prev_scored)]

    # included tokens can be edited in-place
    for (token, idx, is_valid) in included_tokens:
        token_color = (GREEN if is_valid else YELLOW)
        transcription_chars[idx] = token_color + token + CLEARSTYLE

    # omitted tokens must be inserted
    # (NB: an index increment is required as insertion into a list
    # will change the next index we wish to insert at)
    incr = 0
    for (token,idx,_) in omitted_tokens:

        transcription_chars.insert(
            idx+incr,
            f"({RED}{token}{CLEARSTYLE})"
        )
        incr += 1

    evaluation = "".join(transcription_chars)   # combine all tokens for evaluation
    return (score, best_poss_score, evaluation)

def get_results(
    sentence_ipa: str,
    transcription: str,
    scoring_mode: ScoringMode
    ) -> tuple[ list[str], tuple[int, AutoscorerTokens] ]:
    
    if scoring_mode == "preferred-transcription":
        transcription_ipa = str_to_ipa(transcription, True)
        return (transcription_ipa, score_transcription(sentence_ipa, transcription_ipa))

    elif scoring_mode == "best-match":
        poss_results = []
        poss_transcription_ipas = [*map(arpa_to_ipa, to_arpabet_all(transcription))]

        for ipa in poss_transcription_ipas:
            poss_results.append( (ipa, score_transcription(sentence_ipa, ipa)) )

        return max(poss_results, key = lambda n: n[1][0])

def process_inputs(
    inputs: list[tuple[str,str]],
    ws: Worksheet,
    scoring_mode: ScoringMode,
    ideal_ipa: list[str] = ...
    ) -> None:
    
    # utility function to make populating rows of the output file faster
    def fill_row(row: int, values: tuple[str,...]):
        for i, col in enumerate("ABCDEF"):
            ws[f"{col}{row}"] = values[i]

    # write header
    fill_row(
        1, (
            "Source Sentence", "Source IPA",
            "Transcribed Sentence", "Transcribed IPA",
            "Autoscore Evaluation", "Autoscore Score"
            )
    )

    # get a tentative value for a column width to resize to
    col_width = max(max(len(k) for k in  n) for n in inputs)
    col_width = max(20, col_width)                              # ensuring headers are not cropped

    total_score = [0,0]

    (sentences, transcriptions) = [*zip(*inputs)]
    if ideal_ipa is ...:
        sentence_ipas = [*map(str_to_ipa, sentences)]
    else:
        sentence_ipas = [*map(tokenize_ipa, ideal_ipa)]

    # iterate over the target-transcription pairs
    row=2                                               # doing it this way to use row outside of the loop's scope after
    for (sentence, sentence_ipa, transcription) in zip(sentences, sentence_ipas, transcriptions):

        # debug
        # print(f"TARGET: {sentence} | {sentence_ipa}", f"\nSUBJECT: {transcription} | {transcription_ipa}\n\n")

        transcription_ipa, results = get_results(sentence_ipa, transcription, scoring_mode)

        score, best_poss_score, evaluation = output_scoring(                # format the results
            sentence_ipa, transcription_ipa, results
        )

        display_sentence_ipa = "".join(sentence_ipa)
        display_transcription_ipa = "".join(transcription_ipa)

        total_score[0] += score
        total_score[1] += best_poss_score

        formatted_score = f"{score}/{best_poss_score}"
        formatted_evaluation = to_rich_text(evaluation)                     # convert evaluation to rich text

        fill_row(                                                       # write row to output sheet
            row, (
                sentence, display_sentence_ipa,
                transcription, display_transcription_ipa,
                formatted_evaluation, formatted_score
            )
        )

        # if the length of the evaluation is longer than our defined column width, widen it
        col_width = max(col_width, len(str(formatted_evaluation)))
        row += 1

    formatted_total_score = f"{total_score[0]}/{total_score[1]}"
    ws[f"F{row}"] = f"Total Score: {formatted_total_score}"

    # set sheet column dimensions
    dim_holder = dimensions.DimensionHolder(ws)
    for col in "ABCDEF":
        dim_holder[col] = dimensions.ColumnDimension(
            ws, col, width=col_width, bestFit = True
        )

    ws.column_dimensions = dim_holder

def main(
    df: pd.DataFrame = ...,
    out_dir: str = ...,
    out_fn: str = ...,
    ideal_ipa: list[str] = ...,
    scoring_mode: ScoringMode = "preferred-transcription"
    ):

    root = Tk()
    root.withdraw()

    wb = Workbook()
    wb.remove(wb.active)

    if df is None:
        fp = Path(askopenfilename(
            filetypes=[("Excel Files", ".xlsx")]
        ))

        # aborted during filedialog
        if fp == "": return

        if out_dir is None:
            out_dir = fp.parent

        if out_fn is None:
            out_fn = fp.name[:-5] + "_autoscored"

        df = pd.read_excel(fp)

    target = df["Target"]   # get target sentences

    # iterate through subject columns
    for col in df.columns[1:]:
        ws = wb.create_sheet(col)

        inputs: list[tuple[str,str]] = [ *zip(target, df[col]) ]
        process_inputs(inputs, ws, scoring_mode, ideal_ipa)

    outpath = Path(out_dir, f"{out_fn}.xlsx")
    wb.save(outpath)

if __name__ == "__main__":
    main()